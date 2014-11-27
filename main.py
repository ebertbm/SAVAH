#!/usr/bin/python
# -*- coding: utf-8 -*-

import sqlite3 as lite

import sys
import csv
import os
import simplejson
from openpyxl import load_workbook
import datetime
from datetime import date, timedelta
import pdb
import urllib2 



#Define paths
INTERNAL_SERVER = 'WRITE SERVER ROUTE'
URL_FTP_SERVER = "http://www.client.com/data/media/"
ROOT_DIR = os.path.abspath(os.path.dirname(__file__))



#Insert output into a SQLite database
def insertdb(data):
	con = lite.connect(ROOT_DIR+'/havas.db')
	with con:
		cur = con.cursor()
		cur.executemany("INSERT INTO Media VALUES(NULL, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", data)


	print "Output has been generated in the table Media"
	generateJSON()

#Generate a JSON file which will be read by a javascript app that present a chart
def generateJSON():
	con = lite.connect(ROOT_DIR+'/havas.db')
	datadict = []

	with con:
		cur = con.cursor()
		#Query Conversation Rate per Campaign in the Database
		schema = cur.execute("SELECT campaign, ROUND(avg(CR), 2) cr_avg FROM Media group by campaign order by cr_avg desc")
		for x in schema.fetchall():
			media = {
			'campaign' : x[0],
			'cr_avg' : x[1]
			}
			datadict.append(media)


	#print simplejson.dumps(datadict)
	out_file = open(ROOT_DIR+"/media.json",'w')
	simplejson.dump(datadict,out_file, indent=4)     #Generate the JSON file
	print "Generate JSON file"
	print "******Process succesfully executed*****"


	out_file.close()

#Get the CTR
def getctr(clicks, impressions):
	if float(clicks) == 0:
		ctr=0
	elif float(impressions) == 0:
		ctr=clicks
	else:
		ctr = (float(clicks)/float(impressions))*100
	return float(ctr)

#Get the Conversion Rate
def getcr(clicks, sales):
	if float(sales) == 0:
		cr=0
	elif float(clicks) < 1:
		cr=sales*100
	else:
		cr = (float(sales)/float(clicks))*100
	return cr


#Get the Cost Per Acquisition
def getcpa(clicks, cost, sales):
	if float(sales) == 0:
		cpa=cost
	elif float(cost)== 0 or float(clicks) == 0:
		cpa=0
	else:
		cpa = (float(cost)*float(clicks))/float(sales)
	return cpa								
	


#Open the lookup file of campaigns and products
#Create a dictionary that include campaigns and products
def open_mapping():
	f = open(ROOT_DIR+"/campaigns to products.txt") # READ IN THE LOCAL ROOT
	print "Open the file: "+"/campaigns to products.txt"

	campaigns = {} # initialize an empty dictionary
	for line in f:
		campaign, product  = line.split("\t")
		campaigns[campaign] = product.rstrip()
	return campaigns



#Process in a List the data of the excel file.
def process_costfile(ws, today, saturday, sunday):
	data =[]
	datecost = date.today()

	for row in ws.iter_rows():
		if type(row[0].value)==datetime.datetime:
			datecost = row[0].value
			if datecost.strftime("%Y-%m-%d") in (today.strftime("%Y-%m-%d"), saturday.strftime("%Y-%m-%d"), sunday.strftime("%Y-%m-%d")): #Validate if there are records with the specified date
				rowdata = [datecost, row[1].value.rstrip(), row[2].value.rstrip(), row[3].value]
				data.append(rowdata)  #Save rowdata list within another list
	return data

#Open the excel file "costs"
def open_costs():
	wb2 = load_workbook(ROOT_DIR+'/costs.xlsx', use_iterators = True) #ROOT_DIR could be replaced by INTERNAL_SERVER path.
	ws = wb2.get_sheet_by_name('Sheet1')
	print "Open the file: "+"/costs.xlsx"
	today = datetime.date(2014,11,27) #Static Date for testing
	#today = date.today() #Must have a date from today
	data =[]

#Validate which weekday is today to know which records to process
	if today.weekday() in (1,2,3,4):    # This is for Tuesday, Wednesday, Thursday, Friday
		data = process_costfile(ws, today, today, today) #Process today's only records
	elif today.weekday() == 0: #This is Monday
		sunday = date.today() - timedelta(1) 
		saturday = date.today() - timedelta(2)
		data = process_costfile(ws, today, saturday, sunday) #Process the records of monday, sunday and saturday
	else:
		print "It's weekend nobody is working in the Excel file"
	return data


#Main function where join all the data sources
def open_mediadata(today):
	campaigns = open_mapping()
	costs = open_costs()
	data = []

	today='20141127' #Assign a date in this format for test purpose. Comment sentence if want to be today date.
	print "Start process for day: "+today

	try:
		# with open(URL_FTP_SERVER+'/'+today+'/aggregate.csv', 'rU') as csvfile: # Read the file from a FTP server assuming that has OPEN permissions to consult from a HTTP Request
		with open(ROOT_DIR+'/'+today+'aggregate.csv', 'rU') as csvfile:   #Read the .csv file in Local ROOT
			reader = csv.reader(csvfile, delimiter=',', skipinitialspace=True)
			print "Open the file: "+today+'aggregate.csv'
			datecost = date.today()
			print "Join all the files..........."
			for columnData in reader:
				if campaigns.has_key(columnData[1]):
					for rowCost in costs:
						datemedia = datetime.datetime.strptime(columnData[0], "%Y%m%d").strftime('%Y-%m-%d')
						if type(rowCost[0])==datetime.datetime:
							datecost = rowCost[0]
						if datecost.strftime('%Y-%m-%d')==datemedia and rowCost[1]==columnData[1] and rowCost[2]==columnData[2]: #Both files must have equal date, campaign and placement
							#Create a list with all the fields joined
							rowdata = [datemedia, columnData[1], campaigns[columnData[1]], columnData[2], columnData[3], columnData[4], 
							columnData[5], columnData[6], rowCost[3], ("%.3f" % getctr(columnData[4], columnData[5])),("%.3f" % getcr(columnData[4], 
							columnData[6])), ("%.3f" % getcpa(columnData[4], rowCost[3], columnData[6])),
							]
							if rowdata is not None:						
								data.append(rowdata) #Save the data in a list
		insertdb(data) #Insert the data in the DB

	except IOError as e:
		print "I/O error({0}): There is not file for today".format(e.errno, e.strerror)
	except ValueError:
		print "Could not convert"
	except:
		print "Unexpected error:", sys.exc_info()[0]
		raise					



#*****THE SCRIPT START HERE*******#
def main():

	#campaigns = open_mapping()
	today = date.today().strftime("%Y%m%d")
	open_mediadata(today)
	#open_costs()




if __name__ == '__main__':
    main()

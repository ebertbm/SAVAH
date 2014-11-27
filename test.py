from openpyxl import load_workbook
import time
import datetime
from datetime import date, timedelta

wb2 = load_workbook('costs.xlsx', use_iterators = True)
ws = wb2.get_sheet_by_name('Sheet1')
today = date.today()
datecost = date.today()
data = []
rowdata = []

if today.weekday() in (1,2,3,4):
	for row in ws.iter_rows():
		if type(row[0].value)==datetime.datetime:
			datecost = row[0].value
			if today.strftime("%Y-%m-%d") == datecost.strftime("%Y-%m-%d"):
				rowdata = [row[0].value, row[1].value, row[2].value, row[3].value]
				data.append(rowdata)
elif today.weekday() == 0:
	sunday = date.today() - timedelta(1)
	saturday = date.today() - timedelta(2)
	for row in ws.iter_rows():
		if type(row[0].value)==datetime.datetime:
			datecost = row[0].value
			if datecost.strftime("%Y-%m-%d") in (today.strftime("%Y-%m-%d"), saturday.strftime("%Y-%m-%d"), sunday.strftime("%Y-%m-%d")):
				rowdata = [datecost, row[1].value, row[2].value, row[3].value]
				data.append(rowdata)
else:
	print "It's weekend nobody is working in the Excel file"

for x in data:
	print x






"""s = datetime.datetime.strptime("20141207", "%Y%m%d")
print s.strftime('%Y-%m-%d 00:00:00')"""